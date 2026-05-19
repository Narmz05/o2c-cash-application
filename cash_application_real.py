"""
Order to Cash – Cash Application & Invoice Reconciliation System
Author : Narmadhadevi C
Dataset : Real-world O2C Invoice Dataset (Kaggle) — 50,000 records

Description : Applies payments to open invoices, performs reconciliation,
              aging analysis, customer risk scoring, late payment prediction (ML),
              data quality flagging, chart generation, and a complete O2C Excel report.

Outputs:
  - O2C_Full_Report.xlsx     (7 formatted sheets)
  - O2C_Aging_Buckets.png
  - O2C_Collection_Rate.png
  - O2C_Payment_Trend.png
  - O2C_Risk_Distribution.png
"""

import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")          # non-interactive backend — no display needed
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, classification_report
import warnings
warnings.filterwarnings("ignore")

print("=" * 65)
print("  ORDER TO CASH — CASH APPLICATION, RISK SCORING & ML ANALYSIS")
print("=" * 65)

# ─────────────────────────────────────────────────────────────
# 1. LOAD & CLEAN DATA
# ─────────────────────────────────────────────────────────────
df = pd.read_csv("dataset.csv")
print(f"\nDataset Loaded: {len(df):,} records")

def safe_date(val):
    try:
        s = str(int(float(val)))
        return pd.to_datetime(s, format="%Y%m%d", errors="coerce")
    except:
        return pd.NaT

df["due_date"]     = df["due_in_date"].apply(safe_date)
df["create_date"]  = df["document_create_date"].apply(safe_date)
df["clear_date"]   = pd.to_datetime(df["clear_date"], errors="coerce")
df["posting_date"] = pd.to_datetime(df["posting_date"], errors="coerce")
df["amount"]       = pd.to_numeric(df["total_open_amount"], errors="coerce").fillna(0)
df["name_customer"]= df["name_customer"].fillna("Unknown Customer").str.strip()

today = pd.Timestamp.today().normalize()

# ─────────────────────────────────────────────────────────────
# 2. DATA QUALITY FLAGS  (new sheet)
# ─────────────────────────────────────────────────────────────
dq_flags = []

missing_due   = df["due_date"].isna()
missing_clear = df[df["isOpen"] == 0]["clear_date"].isna()
negative_amt  = df["amount"] < 0
unknown_cust  = df["name_customer"].str.upper() == "UNKNOWN CUSTOMER"
future_due    = df["due_date"] > today + pd.Timedelta(days=365 * 5)   # > 5 yrs ahead
zero_amount   = df["amount"] == 0

def flag_row(row):
    issues = []
    if pd.isna(row["due_date"]):
        issues.append("Missing due date")
    if row["isOpen"] == 0 and pd.isna(row["clear_date"]):
        issues.append("Cleared but no clear_date")
    if row["amount"] < 0:
        issues.append("Negative amount")
    if row["amount"] == 0:
        issues.append("Zero amount")
    if str(row["name_customer"]).upper() == "UNKNOWN CUSTOMER":
        issues.append("Unknown customer")
    if not pd.isna(row["due_date"]) and row["due_date"] > today + pd.Timedelta(days=365*5):
        issues.append("Suspicious future due date (>5 yrs)")
    return "; ".join(issues) if issues else "OK"

print("\nRunning data quality checks...")
df["dq_flag"] = df.apply(flag_row, axis=1)
dq_issues = df[df["dq_flag"] != "OK"].copy()

dq_export = dq_issues[[
    "invoice_id", "name_customer", "amount", "due_date", "clear_date", "isOpen", "dq_flag"
]].copy()
dq_export.columns = [
    "Invoice ID", "Customer Name", "Amount ($)", "Due Date", "Clear Date", "Status (0=Paid)", "Issue Flags"
]
dq_export["Due Date"]   = pd.to_datetime(dq_export["Due Date"],   errors="coerce").dt.strftime("%d-%b-%Y")
dq_export["Clear Date"] = pd.to_datetime(dq_export["Clear Date"], errors="coerce").dt.strftime("%d-%b-%Y")

print(f"  Data quality issues found : {len(dq_issues):,} records")
flag_counts = {}
for f in dq_issues["dq_flag"]:
    for issue in f.split("; "):
        flag_counts[issue] = flag_counts.get(issue, 0) + 1
for k, v in flag_counts.items():
    print(f"    {k}: {v:,}")

# ─────────────────────────────────────────────────────────────
# 3. CASH APPLICATION
# ─────────────────────────────────────────────────────────────
# Exclude unknown customers from operational analysis (keep for DQ sheet)
clean_df  = df[df["name_customer"].str.upper() != "UNKNOWN CUSTOMER"].copy()
paid_df   = clean_df[clean_df["isOpen"] == 0].copy()
open_df   = clean_df[clean_df["isOpen"] == 1].copy()

print(f"\nInvoice Summary (excl. unknown customers)")
print(f"  Total Invoices     : {len(clean_df):,}")
print(f"  Paid / Cleared     : {len(paid_df):,}")
print(f"  Open / Outstanding : {len(open_df):,}")

def cash_app_status(row):
    if pd.isna(row["clear_date"]) or pd.isna(row["due_date"]):
        return "Cleared - Date Unknown"
    if row["clear_date"] <= row["due_date"]:
        return "Paid - On Time"
    else:
        days_late = (row["clear_date"] - row["due_date"]).days
        if days_late <= 15:
            return "Paid - Slightly Late (<=15 days)"
        elif days_late <= 30:
            return "Paid - Late (16-30 days)"
        else:
            return "Paid - Very Late (>30 days)"

paid_df["application_status"] = paid_df.apply(cash_app_status, axis=1)
paid_df["days_to_pay"]        = (paid_df["clear_date"] - paid_df["due_date"]).dt.days

# ─────────────────────────────────────────────────────────────
# 4. AGING ANALYSIS
# ─────────────────────────────────────────────────────────────
def aging_bucket(row):
    if pd.isna(row["due_date"]):
        return "Unknown"
    days = (today - row["due_date"]).days
    if days <= 0:
        return "Current (Not Yet Due)"
    elif days <= 30:
        return "1-30 Days Overdue"
    elif days <= 60:
        return "31-60 Days Overdue"
    elif days <= 90:
        return "61-90 Days Overdue"
    else:
        return "90+ Days Overdue"

open_df["aging_bucket"]  = open_df.apply(aging_bucket, axis=1)
open_df["days_overdue"]  = (today - open_df["due_date"]).dt.days.clip(lower=0)

# ─────────────────────────────────────────────────────────────
# 5. CUSTOMER RISK SCORING  (new column on open_df)
# ─────────────────────────────────────────────────────────────
# Formula: risk = (days_overdue_norm × 0.4) + (amount_norm × 0.6)
# Normalize each to 0-1, then combine

max_days   = open_df["days_overdue"].max() or 1
max_amount = open_df["amount"].max() or 1

open_df["days_overdue_norm"] = open_df["days_overdue"] / max_days
open_df["amount_norm"]       = open_df["amount"] / max_amount
open_df["risk_score"]        = (
    open_df["days_overdue_norm"] * 0.4 +
    open_df["amount_norm"]       * 0.6
).round(4)

def risk_label(score):
    if score >= 0.6:
        return "HIGH"
    elif score >= 0.3:
        return "MEDIUM"
    else:
        return "LOW"

open_df["risk_level"] = open_df["risk_score"].apply(risk_label)

risk_summary = open_df["risk_level"].value_counts()
print(f"\nCustomer Risk Distribution (Open Invoices):")
for level in ["HIGH", "MEDIUM", "LOW"]:
    count = risk_summary.get(level, 0)
    amount = open_df[open_df["risk_level"] == level]["amount"].sum()
    print(f"  {level:6s}: {count:6,} invoices   ${amount:>15,.2f}")

# ─────────────────────────────────────────────────────────────
# 6. ML — LATE PAYMENT PREDICTION
# ─────────────────────────────────────────────────────────────
print("\nTraining ML model: Late Payment Prediction...")

ml_df = paid_df.dropna(subset=["due_date", "clear_date", "business_code"]).copy()
ml_df["is_late"]     = (ml_df["days_to_pay"] > 0).astype(int)
ml_df["invoice_age"] = (ml_df["due_date"] - ml_df["create_date"]).dt.days.fillna(30)

le_biz = LabelEncoder()
ml_df["biz_code_enc"] = le_biz.fit_transform(ml_df["business_code"].astype(str))

features = ["biz_code_enc", "amount", "invoice_age"]
X = ml_df[features].fillna(0)
y = ml_df["is_late"]

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42, stratify=y
)

scaler = StandardScaler()
X_train_s = scaler.fit_transform(X_train)
X_test_s  = scaler.transform(X_test)

model = LogisticRegression(max_iter=500, random_state=42)
model.fit(X_train_s, y_train)
y_pred = model.predict(X_test_s)

accuracy = accuracy_score(y_test, y_pred)
print(f"  Model Accuracy     : {accuracy:.1%}")
report   = classification_report(y_test, y_pred, target_names=["On Time", "Late"])
print(report)

# Apply model predictions to open invoices (to flag which are likely to pay late)
open_ml = open_df.dropna(subset=["business_code"]).copy()
open_ml["invoice_age"] = (open_ml["due_date"] - open_ml["create_date"]).dt.days.fillna(30)
open_ml["biz_code_enc"] = open_ml["business_code"].astype(str).map(
    lambda x: le_biz.transform([x])[0] if x in le_biz.classes_ else -1
)
valid_open = open_ml[open_ml["biz_code_enc"] >= 0].copy()
X_open = valid_open[features].fillna(0)
X_open_s = scaler.transform(X_open)
valid_open["late_prob"]       = model.predict_proba(X_open_s)[:, 1].round(3)
valid_open["payment_prediction"] = valid_open["late_prob"].apply(
    lambda p: "Likely Late" if p >= 0.5 else "Likely On Time"
)
open_df = open_df.merge(
    valid_open[["invoice_id", "late_prob", "payment_prediction"]],
    on="invoice_id", how="left"
)
open_df["payment_prediction"] = open_df["payment_prediction"].fillna("Insufficient Data")

# ─────────────────────────────────────────────────────────────
# 7. SUMMARY STATS
# ─────────────────────────────────────────────────────────────
total_invoiced  = clean_df["amount"].sum()
total_collected = paid_df["amount"].sum()
total_open_amt  = open_df["amount"].sum()
collection_rate = (total_collected / total_invoiced * 100) if total_invoiced > 0 else 0
avg_days_late   = paid_df["days_to_pay"].mean()
dso             = (total_open_amt / total_invoiced * 365) if total_invoiced > 0 else 0

print(f"\nFinancial Summary")
print(f"  Total Invoiced     : ${total_invoiced:>15,.2f}")
print(f"  Total Collected    : ${total_collected:>15,.2f}")
print(f"  Total Outstanding  : ${total_open_amt:>15,.2f}")
print(f"  Collection Rate    : {collection_rate:>14.1f}%")
print(f"  Avg Days to Pay    : {avg_days_late:>14.1f} days")
print(f"  Estimated DSO      : {dso:>14.1f} days")

print(f"\nCash Application Status Breakdown:")
print(paid_df["application_status"].value_counts().to_string())

print(f"\nAR Aging Breakdown (Open Invoices):")
aging_summary = open_df.groupby("aging_bucket")["amount"].agg(["count","sum"])
aging_summary.columns = ["Count","Amount ($)"]
aging_summary["Amount ($)"] = aging_summary["Amount ($)"].map("${:,.2f}".format)
print(aging_summary.to_string())

# ─────────────────────────────────────────────────────────────
# 8. CHARTS
# ─────────────────────────────────────────────────────────────
BUCKET_ORDER = [
    "Current (Not Yet Due)", "1-30 Days Overdue",
    "31-60 Days Overdue",    "61-90 Days Overdue",
    "90+ Days Overdue",      "Unknown"
]
BUCKET_COLORS = ["#2ecc71","#f1c40f","#e67e22","#e74c3c","#8e44ad","#95a5a6"]

# Chart 1 — AR Aging Buckets (bar)
aging_bar = open_df.groupby("aging_bucket")["amount"].sum().reindex(
    [b for b in BUCKET_ORDER if b in open_df["aging_bucket"].unique()]
).dropna()

fig, ax = plt.subplots(figsize=(10, 5))
colors_used = [BUCKET_COLORS[BUCKET_ORDER.index(b)] for b in aging_bar.index]
bars = ax.bar(aging_bar.index, aging_bar.values / 1e6, color=colors_used, edgecolor="white", linewidth=0.8)
ax.set_title("AR Aging — Outstanding by Bucket", fontsize=14, fontweight="bold", pad=14)
ax.set_xlabel("Aging Bucket")
ax.set_ylabel("Outstanding Amount (USD Millions)")
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x, _: f"${x:.1f}M"))
for bar in bars:
    h = bar.get_height()
    ax.text(bar.get_x() + bar.get_width()/2, h + 0.2, f"${h:.1f}M",
            ha="center", va="bottom", fontsize=9, fontweight="bold")
plt.xticks(rotation=20, ha="right")
plt.tight_layout()
plt.savefig("O2C_Aging_Buckets.png", dpi=150)
plt.close()
print("\nChart saved: O2C_Aging_Buckets.png")

# Chart 2 — Collection Rate (pie / donut)
collected_pct  = collection_rate
outstanding_pct = 100 - collection_rate
fig, ax = plt.subplots(figsize=(6, 6))
wedges, texts, autotexts = ax.pie(
    [collected_pct, outstanding_pct],
    labels=["Collected", "Outstanding"],
    colors=["#27ae60","#e74c3c"],
    autopct="%1.1f%%",
    startangle=90,
    wedgeprops=dict(width=0.55),
    textprops=dict(fontsize=12)
)
autotexts[0].set_fontweight("bold")
autotexts[1].set_fontweight("bold")
ax.set_title(f"Collection Rate: {collection_rate:.1f}%", fontsize=14, fontweight="bold")
plt.tight_layout()
plt.savefig("O2C_Collection_Rate.png", dpi=150)
plt.close()
print("Chart saved: O2C_Collection_Rate.png")

# Chart 3 — Monthly Payment Trend
paid_trend = paid_df.dropna(subset=["clear_date"]).copy()
paid_trend["month"] = paid_trend["clear_date"].dt.to_period("M")
monthly = paid_trend.groupby("month")["amount"].sum().reset_index()
monthly["month_str"] = monthly["month"].astype(str)
# Keep last 24 months for readability
monthly = monthly.tail(24)

fig, ax = plt.subplots(figsize=(12, 5))
ax.plot(monthly["month_str"], monthly["amount"] / 1e6,
        marker="o", color="#2980b9", linewidth=2, markersize=5)
ax.fill_between(range(len(monthly)), monthly["amount"] / 1e6, alpha=0.12, color="#2980b9")
ax.set_title("Monthly Collections Trend (Last 24 Months)", fontsize=14, fontweight="bold", pad=14)
ax.set_xlabel("Month")
ax.set_ylabel("Amount Collected (USD Millions)")
ax.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x, _: f"${x:.1f}M"))
ax.set_xticks(range(len(monthly)))
ax.set_xticklabels(monthly["month_str"], rotation=45, ha="right", fontsize=8)
ax.grid(axis="y", linestyle="--", alpha=0.5)
plt.tight_layout()
plt.savefig("O2C_Payment_Trend.png", dpi=150)
plt.close()
print("Chart saved: O2C_Payment_Trend.png")

# Chart 4 — Risk Distribution (horizontal bar)
risk_counts  = open_df["risk_level"].value_counts().reindex(["HIGH","MEDIUM","LOW"]).fillna(0)
risk_amounts = open_df.groupby("risk_level")["amount"].sum().reindex(["HIGH","MEDIUM","LOW"]).fillna(0)

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))
risk_colors = ["#e74c3c","#e67e22","#27ae60"]

ax1.barh(risk_counts.index, risk_counts.values, color=risk_colors, edgecolor="white")
ax1.set_title("Invoice Count by Risk Level", fontweight="bold")
ax1.set_xlabel("Number of Invoices")
for i, v in enumerate(risk_counts.values):
    ax1.text(v + 10, i, f"{int(v):,}", va="center", fontsize=10)

ax2.barh(risk_amounts.index, risk_amounts.values / 1e6, color=risk_colors, edgecolor="white")
ax2.set_title("Outstanding Amount by Risk Level", fontweight="bold")
ax2.set_xlabel("Amount (USD Millions)")
ax2.xaxis.set_major_formatter(mtick.FuncFormatter(lambda x, _: f"${x:.1f}M"))
for i, v in enumerate(risk_amounts.values):
    ax2.text(v/1e6 + 0.05, i, f"${v/1e6:.1f}M", va="center", fontsize=10)

plt.suptitle("Customer Risk Scoring — Open Invoices", fontsize=13, fontweight="bold", y=1.02)
plt.tight_layout()
plt.savefig("O2C_Risk_Distribution.png", dpi=150, bbox_inches="tight")
plt.close()
print("Chart saved: O2C_Risk_Distribution.png")

# ─────────────────────────────────────────────────────────────
# 9. PREPARE EXPORT DATA
# ─────────────────────────────────────────────────────────────
cash_app_export = paid_df[[
    "invoice_id","name_customer","business_code","invoice_currency",
    "amount","create_date","due_date","clear_date",
    "days_to_pay","application_status"
]].copy()
cash_app_export.columns = [
    "Invoice ID","Customer Name","Business Code","Currency",
    "Invoice Amount ($)","Invoice Date","Due Date","Payment Date",
    "Days to Pay","Application Status"
]
for col in ["Invoice Date","Due Date","Payment Date"]:
    cash_app_export[col] = pd.to_datetime(cash_app_export[col]).dt.strftime("%d-%b-%Y")

aging_export = open_df[[
    "invoice_id","name_customer","business_code","invoice_currency",
    "amount","due_date","days_overdue","aging_bucket",
    "risk_score","risk_level","payment_prediction"
]].copy()
aging_export.columns = [
    "Invoice ID","Customer Name","Business Code","Currency",
    "Outstanding Amount ($)","Due Date","Days Overdue","Aging Bucket",
    "Risk Score","Risk Level","ML Payment Prediction"
]
aging_export["Due Date"] = pd.to_datetime(aging_export["Due Date"]).dt.strftime("%d-%b-%Y")

aging_pivot = open_df.groupby("aging_bucket")["amount"].agg(
    Count="count", Total_Amount="sum"
).reset_index()
aging_pivot.columns = ["Aging Bucket","Invoice Count","Total Outstanding ($)"]

top_customers = (
    open_df.groupby("name_customer")["amount"]
    .sum().sort_values(ascending=False).head(10).reset_index()
)
top_customers.columns = ["Customer Name","Outstanding Amount ($)"]
top_cust_export = top_customers.copy()

summary_df = pd.DataFrame({
    "Metric": [
        "Total Invoices","Paid Invoices","Open Invoices",
        "Total Invoiced ($)","Total Collected ($)",
        "Total Outstanding ($)","Collection Rate (%)","Avg Days to Pay",
        "Estimated DSO (days)","ML Model Accuracy",
        "Data Quality Issues","HIGH Risk Invoices","MEDIUM Risk Invoices","LOW Risk Invoices"
    ],
    "Value": [
        f"{len(clean_df):,}", f"{len(paid_df):,}", f"{len(open_df):,}",
        f"${total_invoiced:,.2f}", f"${total_collected:,.2f}",
        f"${total_open_amt:,.2f}", f"{collection_rate:.1f}%",
        f"{avg_days_late:.1f} days", f"{dso:.1f} days",
        f"{accuracy:.1%}",
        f"{len(dq_issues):,}",
        f"{risk_summary.get('HIGH', 0):,}",
        f"{risk_summary.get('MEDIUM', 0):,}",
        f"{risk_summary.get('LOW', 0):,}"
    ]
})

# ─────────────────────────────────────────────────────────────
# 10. EXPORT TO EXCEL  (7 sheets)
# ─────────────────────────────────────────────────────────────
output_file = "O2C_Full_Report.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    summary_df.to_excel(writer,   sheet_name="Executive Summary",    index=False)
    cash_app_export.to_excel(writer, sheet_name="Cash Application",  index=False)
    aging_export.to_excel(writer, sheet_name="AR Aging Detail",      index=False)
    aging_pivot.to_excel(writer,  sheet_name="Aging Buckets",        index=False)
    top_cust_export.to_excel(writer, sheet_name="Top 10 Outstanding",index=False)
    dq_export.to_excel(writer,    sheet_name="Data Quality Issues",  index=False)

    # ML Results sheet
    ml_results = pd.DataFrame({
        "Metric": ["Model Type","Features Used","Training Records","Test Records","Accuracy","Notes"],
        "Value": [
            "Logistic Regression",
            "Business Code, Invoice Amount, Invoice Age (days)",
            f"{len(X_train):,}",
            f"{len(X_test):,}",
            f"{accuracy:.1%}",
            "Predicts whether an invoice will be paid late. Applied to all open invoices."
        ]
    })
    ml_results.to_excel(writer, sheet_name="ML Model Results", index=False)

# ─────────────────────────────────────────────────────────────
# 11. FORMAT EXCEL
# ─────────────────────────────────────────────────────────────
STATUS_COLORS = {
    "Paid - On Time"                  : "C6EFCE",
    "Paid - Slightly Late (<=15 days)": "FFEB9C",
    "Paid - Late (16-30 days)"        : "FFCC99",
    "Paid - Very Late (>30 days)"     : "FFC7CE",
    "Cleared - Date Unknown"          : "DDEBF7",
    "Current (Not Yet Due)"           : "C6EFCE",
    "1-30 Days Overdue"               : "FFEB9C",
    "31-60 Days Overdue"              : "FFCC99",
    "61-90 Days Overdue"              : "FFC7CE",
    "90+ Days Overdue"                : "FF4444",
    "HIGH"                            : "FFC7CE",
    "MEDIUM"                          : "FFEB9C",
    "LOW"                             : "C6EFCE",
    "Likely Late"                     : "FFC7CE",
    "Likely On Time"                  : "C6EFCE",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER

    # Identify color-coded columns
    color_cols = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value in ("Application Status", "Aging Bucket", "Risk Level", "ML Payment Prediction"):
            color_cols[idx] = cell.value

    # Data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center")

        for col_idx, col_name in color_cols.items():
            val   = row[col_idx - 1].value
            color = STATUS_COLORS.get(str(val), "FFFFFF")
            if color != "FFFFFF":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=color)
                break   # only one color per row

    # Column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 35)

    ws.freeze_panes = "A2"

wb.save(output_file)

print(f"\n{'='*65}")
print(f"  REPORT COMPLETE")
print(f"{'='*65}")
print(f"  Excel  : {output_file}  (7 sheets)")
print(f"  Charts : O2C_Aging_Buckets.png")
print(f"           O2C_Collection_Rate.png")
print(f"           O2C_Payment_Trend.png")
print(f"           O2C_Risk_Distribution.png")
print(f"{'='*65}")